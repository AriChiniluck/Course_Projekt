"""Tools for the Telekom business customer support multi-agent system.

Provides four tool groups:
  1. Customer data search  — reads *.xlsx from the Data/ directory.
  2. Knowledge base search — semantic+BM25 search over internal FAQ/tariff docs.
  3. Tariff lookup         — fetches the operator tariff archive page.
  4. Escalation email      — sends SMTP e-mail and persists the record to DB.
"""
from __future__ import annotations

import smtplib
import ssl
from datetime import date
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email import encoders
from pathlib import Path
from typing import Annotated

import openpyxl
import trafilatura
from ddgs import DDGS
from langchain_core.runnables import RunnableConfig
from langchain_core.tools import InjectedToolArg, tool

from config import settings
from retriever import get_retriever
from user_memory import format_chat_transcript, save_escalation, save_identified_customer

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Customer billing data — lives inside the project under data/customers/
_DATA_DIR = Path(__file__).resolve().parent / "data" / "customers"

_TELEKOM_TARIFF_URL = settings.telekom_tariff_url

# Column header name sets — lowercase, Ukrainian only
_COL_ACCOUNT   = {"особовий рахунок"}
_COL_NAME      = {"назва клієнта", "найменування клієнта",
                  "назва підприємства", "найменування"}
_COL_ID        = {"ідентифікатор", "єдрпоу", "код єдрпоу"}
_COL_PHONE     = {"телефон", "номер телефону", "phone", "номер"}
_COL_TARIFF    = {"тарифний план", "тариф", "назва тарифу"}
_COL_TOTAL     = {"сума всього", "загальна сума", "разом", "сума"}

# Keywords that identify the base subscription fee column — must be excluded from
# paid_services (extra charges) and reported separately.
_ABONPLATA_KW  = ("абонентська плата", "абонплата", "щомісячна плата")


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _normalize_phone(raw: str) -> str:
    """Повертає лише цифри, видаляє український префікс (+38 / 380).
    IoT SIM-ідентифікатори 7-значні, не повні мобільні номери.
    """
    digits = "".join(c for c in str(raw or "") if c.isdigit())
    # Прибраємо префікс +380 або +38 щоб порівнювати суфікс-ідентифікатори без префіксу.
    if digits.startswith("380"):
        digits = digits[3:]
    elif digits.startswith("38") and len(digits) > 10:
        digits = digits[2:]
    return digits


def _col_idx(headers: list[str], name_set: set[str]) -> int | None:
    for i, h in enumerate(headers):
        if str(h).strip().lower() in name_set:
            return i
    return None


def _cell(row: tuple, idx: int) -> str:
    try:
        v = row[idx]
        return str(v).strip() if v is not None else ""
    except IndexError:
        return ""


# Column suffix markers
_SUM_MARKERS = ("|Сума", "|Сумма")
_QTY_MARKERS = ("|Кількість", "|Кол-во")
_VOL_MARKERS = ("|Обсяг", "|Объём", "|Объем")

# Keywords for unit detection (lowercase)
_KW_SMS  = ("sms", "mms", "повідомлення", "viber", "промо-розсилка")
_KW_DATA = ("gprs", "edge", "передача даних", "інтернет", "трафік")
_KW_CALL = ("виклик", "дзвінок", "роумінг", "переадресов")
_KW_SIM  = ("абонентська плата",)


def _qty_unit(base_lower: str) -> str:
    """Return the appropriate unit label for a |Кількість column."""
    if any(k in base_lower for k in _KW_SMS):
        return " шт."
    if any(k in base_lower for k in _KW_CALL):
        return " хв."
    if any(k in base_lower for k in _KW_SIM):
        return " SIM"
    return ""  # unknown — show raw number


def _vol_unit(base_lower: str) -> str:
    """Return the appropriate unit label for a |Обсяг column."""
    if any(k in base_lower for k in _KW_DATA):
        return " МБ"
    if any(k in base_lower for k in _KW_CALL):
        return " хв."
    return ""  # unknown


def _is_abonplata(name: str) -> bool:
    """Return True if this service column header is the base subscription fee."""
    n = name.lower()
    return any(kw in n for kw in _ABONPLATA_KW)


def _extract_services(row: tuple, headers: list[str], start_col: int = 6) -> dict[str, str]:
    """Collect non-zero service cost columns as {service_name: 'X грн (detail)'}.
    Excludes the base subscription fee (абонентська плата) — returned separately.
    """
    # Build index: service_base_name → {type: col_index}
    service_cols: dict[str, dict[str, int]] = {}
    for i in range(start_col, len(headers)):
        h = headers[i]
        if not h:
            continue
        for m in _SUM_MARKERS:
            if m in h:
                service_cols.setdefault(h.replace(m, "").strip(), {})["sum"] = i
                break
        for m in _QTY_MARKERS:
            if m in h:
                service_cols.setdefault(h.replace(m, "").strip(), {})["qty"] = i
                break
        for m in _VOL_MARKERS:
            if m in h:
                service_cols.setdefault(h.replace(m, "").strip(), {})["vol"] = i
                break

    services: dict[str, str] = {}
    for base, cols in service_cols.items():
        if "sum" not in cols:
            continue
        # Skip абонентська плата — it is NOT an extra charge
        if _is_abonplata(base):
            continue
        cost = _cell(row, cols["sum"])
        if not cost or cost in {"0", "0,0", "0.0", "None", ""}:
            continue

        base_lower = base.lower()
        details = []

        qty = _cell(row, cols["qty"]) if "qty" in cols else ""
        if qty and qty not in {"0", "0,0", "0.0", "None", ""}:
            details.append(f"{qty}{_qty_unit(base_lower)}")

        vol = _cell(row, cols["vol"]) if "vol" in cols else ""
        if vol and vol not in {"0", "0,0", "0.0", "None", ""}:
            # show vol only when qty is absent or unit differs
            vol_unit = _vol_unit(base_lower)
            qty_unit = _qty_unit(base_lower)
            if not details or vol_unit != qty_unit:
                details.append(f"{vol}{vol_unit}")

        suffix = f" ({', '.join(details)})" if details else ""
        services[base] = f"{cost} грн{suffix}"
    return services


_ZERO = {"0", "0,0", "0.0", "None", ""}


def _extract_included_usage(row: tuple, headers: list[str], start_col: int = 6) -> dict[str, str]:
    """Return services used within the tariff (cost=0 but qty or vol > 0)."""
    service_cols: dict[str, dict[str, int]] = {}
    for i in range(start_col, len(headers)):
        h = headers[i]
        if not h:
            continue
        for m in _SUM_MARKERS:
            if m in h:
                service_cols.setdefault(h.replace(m, "").strip(), {})["sum"] = i
                break
        for m in _QTY_MARKERS:
            if m in h:
                service_cols.setdefault(h.replace(m, "").strip(), {})["qty"] = i
                break
        for m in _VOL_MARKERS:
            if m in h:
                service_cols.setdefault(h.replace(m, "").strip(), {})["vol"] = i
                break

    usage: dict[str, str] = {}
    for base, cols in service_cols.items():
        # Skip if there's no sum column or cost is non-zero (already in paid_services)
        if "sum" not in cols:
            continue
        cost = _cell(row, cols["sum"])
        if cost and cost not in _ZERO:
            continue  # paid service — handled elsewhere

        base_lower = base.lower()
        details = []

        qty = _cell(row, cols["qty"]) if "qty" in cols else ""
        if qty and qty not in _ZERO:
            details.append(f"{qty}{_qty_unit(base_lower)}")

        vol = _cell(row, cols["vol"]) if "vol" in cols else ""
        if vol and vol not in _ZERO:
            vol_unit = _vol_unit(base_lower)
            qty_unit = _qty_unit(base_lower)
            if not details or vol_unit != qty_unit:
                details.append(f"{vol}{vol_unit}")

        if details:  # only include if actually used
            usage[base] = ", ".join(details)
    return usage


def _search_xlsx(path: Path, query: str) -> list[dict]:
    """Return all rows in one file that match the query string."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        all_rows: list[tuple] = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception:
        return []

    if len(all_rows) < 2:
        return []

    headers = [str(h).strip() if h is not None else "" for h in all_rows[0]]

    # Map column names → indices (with positional fallbacks for non-standard layouts)
    i_account = _col_idx(headers, _COL_ACCOUNT) if _col_idx(headers, _COL_ACCOUNT) is not None else 0
    i_name    = _col_idx(headers, _COL_NAME)    if _col_idx(headers, _COL_NAME)    is not None else 1
    i_id      = _col_idx(headers, _COL_ID)      if _col_idx(headers, _COL_ID)      is not None else 2
    i_phone   = _col_idx(headers, _COL_PHONE)   if _col_idx(headers, _COL_PHONE)   is not None else 3
    i_tariff  = _col_idx(headers, _COL_TARIFF)  if _col_idx(headers, _COL_TARIFF)  is not None else 4
    i_total   = _col_idx(headers, _COL_TOTAL)   if _col_idx(headers, _COL_TOTAL)   is not None else 5

    # Service detail columns start right after total; None/empty headers are skipped in _extract_services
    _service_start = i_total + 1

    q_digits = _normalize_phone(query)
    q_lower  = query.strip().lower()

    results: list[dict] = []
    seen: set[str] = set()

    for row in all_rows[1:]:
        if not row or all(v is None for v in row):
            continue

        account    = _cell(row, i_account)
        name       = _cell(row, i_name)
        identifier = _cell(row, i_id)
        phone      = _cell(row, i_phone)
        tariff     = _cell(row, i_tariff)
        total      = _cell(row, i_total)

        phone_digits = _normalize_phone(phone)

        # If primary phone column looks suspicious (< 6 digits), try nearby columns
        # This handles files where phone is at idx 4 instead of 3
        if len(phone_digits) < 6:
            for alt_idx in range(max(0, i_phone - 1), min(i_phone + 4, len(row))):
                alt = _normalize_phone(_cell(row, alt_idx))
                if len(alt) >= 7:
                    phone_digits = alt
                    phone = _cell(row, alt_idx)
                    break

        # Also check a few extra positions for phone-like values in unusual layouts
        extra_phones = [_normalize_phone(_cell(row, j)) for j in range(3, min(8, len(row)))]

        matched = (
            (q_digits and phone_digits and (q_digits in phone_digits or phone_digits in q_digits)) or
            (q_digits and any(q_digits in ep or ep in q_digits for ep in extra_phones if len(ep) >= 7)) or
            (q_lower and account.lower() == q_lower) or
            (q_lower and identifier.lower() == q_lower) or
            (q_lower and len(q_lower) >= 4 and q_lower in name.lower())
        )

        if not matched:
            continue

        key = f"{account}|{phone}"
        if key in seen:
            continue
        seen.add(key)

        # Extract subscription fee separately from service columns
        subscription_fee = ""
        for i in range(_service_start, len(headers)):
            h = headers[i]
            if h and any(m in h for m in _SUM_MARKERS) and _is_abonplata(h.split("|")[0]):
                subscription_fee = _cell(row, i)
                break

        results.append({
            "account_number":   account,
            "client_name":      name,
            "identifier":       identifier,
            "phone":            phone,
            "tariff_name":      tariff,
            "total_amount":     total,
            "subscription_fee": subscription_fee,
            "source_file":      path.name,
            "paid_services":    _extract_services(row, headers, start_col=_service_start),
            "included_usage":   _extract_included_usage(row, headers, start_col=_service_start),
        })

    return results


@tool
def search_customer(query: str) -> str:
    """Search for a customer in billing data files.

    Use this as soon as the customer provides any identifier:
    phone number, account number (особовий рахунок), ЄДРПОУ, or company name.

    Returns matching records with: client name, account, tariff plan,
    total charges for the billing period, and a breakdown of paid services.
    """
    # Мінімальна довжина ідентифікатора — захист від подачі одних
    # цифр (напр. "0" або "1"), що не є реальним ідентифікатором.
    q = str(query or "").strip()
    if not q or len(q) < 3:
        return (
            "Запит занадто короткий. Вкажіть номер телефону, "
            "особовий рахунок або ÄДРПОУ."
        )

    data_dir = _DATA_DIR
    if not data_dir.exists():
        return f"Директорія з даними не знайдена: {data_dir}"

    all_results: list[dict] = []
    # ~$ префікс — тимчасовий файл Excel (відкритий в редакторі), пропускаємо.
    for xlsx_path in sorted(data_dir.glob("*.xlsx")):
        if xlsx_path.name.startswith("~$"):
            continue
        all_results.extend(_search_xlsx(xlsx_path, q))

    if not all_results:
        return (
            f"Клієнта за запитом «{q}» не знайдено у жодному файлі. "
            "Уточніть дані або спробуйте інший ідентифікатор."
        )

    lines = [f"Знайдено {len(all_results)} запис(ів) за запитом «{q}»:"]
    for r in all_results:
        lines += [
            "",
            f"Клієнт:            {r['client_name']}",
            f"Особовий рахунок:  {r['account_number']}",
            f"ЄДРПОУ:            {r['identifier']}",
            f"Телефон:           {r['phone']}",
            f"Тарифний план:     {r['tariff_name']}",
            f"Абонентська плата: {r['subscription_fee']} грн" if r['subscription_fee'] else f"Загальна сума:     {r['total_amount']} грн",
            f"Загальна сума:     {r['total_amount']} грн",
            f"Файл даних:        {r['source_file']}",
        ]
        if r["paid_services"]:
            lines.append("Позатарифні нарахування (понад абонплату):")
            for svc, amt in list(r["paid_services"].items())[:12]:
                lines.append(f"  • {svc}: {amt}")
            if len(r["paid_services"]) > 12:
                lines.append(f"  … ще {len(r['paid_services']) - 12} позицій")
        else:
            lines.append("Позатарифні нарахування: ВІДСУТНІ (загальна сума = лише абонплата)")
        if r["included_usage"]:
            lines.append("Використання в межах тарифу (0 грн, входить в абонплату):")
            for svc, detail in list(r["included_usage"].items())[:12]:
                lines.append(f"  • {svc}: {detail}")
            if len(r["included_usage"]) > 12:
                lines.append(f"  … ще {len(r['included_usage']) - 12} позицій")
        else:
            lines.append("Використання в межах тарифу: немає даних")
        if not r["paid_services"] and not r["included_usage"]:
            lines.append("Номер не використовував жодних послуг у цьому періоді.")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Tool 1c — SIM-level breakdown: all numbers with non-zero extra charges
# ---------------------------------------------------------------------------

@tool
def get_customer_sims(account_number: str) -> str:
    """Return all SIM numbers for a customer with non-zero extra charges.

    Use this when the customer asks:
    - "Which numbers have extra charges?"
    - "Show me numbers with costs above the subscription fee"
    - "Which SIMs are generating additional costs?"

    Call search_customer first to get the account_number, then call this tool.

    Args:
        account_number: The customer's account number (особовий рахунок),
                        e.g. "774466377". Returned by search_customer.
    """
    acc = str(account_number or "").strip()
    if not acc or len(acc) < 3:
        return "Вкажіть номер особового рахунку (отриманий через search_customer)."

    data_dir = _DATA_DIR
    if not data_dir.exists():
        return f"Директорія з даними не знайдена: {data_dir}"

    all_sims: list[dict] = []
    source_file = ""

    for xlsx_path in sorted(data_dir.glob("*.xlsx")):
        if xlsx_path.name.startswith("~$"):
            continue
        try:
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb.active
            all_rows: list[tuple] = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception:
            continue

        if len(all_rows) < 2:
            continue

        headers = [str(h).strip() if h is not None else "" for h in all_rows[0]]
        i_account = _col_idx(headers, _COL_ACCOUNT) if _col_idx(headers, _COL_ACCOUNT) is not None else 0
        i_name    = _col_idx(headers, _COL_NAME)    if _col_idx(headers, _COL_NAME)    is not None else 1
        i_phone   = _col_idx(headers, _COL_PHONE)   if _col_idx(headers, _COL_PHONE)   is not None else 3
        i_tariff  = _col_idx(headers, _COL_TARIFF)  if _col_idx(headers, _COL_TARIFF)  is not None else 4
        i_total   = _col_idx(headers, _COL_TOTAL)   if _col_idx(headers, _COL_TOTAL)   is not None else 5
        svc_start = i_total + 1

        for row in all_rows[1:]:
            if not row or all(v is None for v in row):
                continue
            if _cell(row, i_account).strip() != acc:
                continue

            phone = _cell(row, i_phone)
            if not phone:
                continue

            paid = _extract_services(row, headers, start_col=svc_start)
            if not paid:
                continue  # no extra charges — skip

            all_sims.append({
                "phone":    phone,
                "tariff":   _cell(row, i_tariff),
                "total":    _cell(row, i_total),
                "services": paid,
            })
            if not source_file:
                source_file = xlsx_path.name
                client_name = _cell(row, i_name)

    if not all_sims:
        return (
            f"Для рахунку «{acc}» не знайдено SIM-карток з позатарифними витратами. "
            "Перевірте номер рахунку або скористайтесь search_customer."
        )

    lines = [
        f"Клієнт: {client_name}  (рахунок {acc})",
        f"SIM-номери з позатарифними нарахуваннями: {len(all_sims)}",
        "",
        "Пояснення полів:",
        "  абонплата — базова щомісячна плата за тариф (включена в план)",
        "  позатарифні нарахування — додаткові послуги понад абонплату",
        "",
    ]
    for sim in all_sims:
        lines.append(f"📱 {sim['phone']}  |  тариф: {sim['tariff']}  |  абонплата: {sim['total']} грн")
        lines.append(f"   Позатарифні нарахування:")
        for svc, amt in sim["services"].items():
            lines.append(f"    • {svc}: {amt}")
        lines.append("")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Tool 1b — Internal knowledge base search (FAQ + tariff docs)
# ---------------------------------------------------------------------------

@tool
def search_telekom_kb(question: str) -> str:
    """Search the internal Telekom FAQ and tariff knowledge base.

    Use for:
    - Explaining tariff conditions (IoT 17 / IoT 45 / IoT 70)
    - Answering billing questions (what is a charge, what is included)
    - Explaining why numbers get blocked and how to unblock
    - Credit limit questions
    - SIM management procedures

    Prefer this over lookup_tariff_telekom for general questions.
    Use lookup_tariff_telekom only when you need live data from the operator website.
    """
    q = str(question or "").strip()
    if not q or len(q) < 3:
        return "Запит занадто короткий."

    try:
        retriever = get_retriever()
    except FileNotFoundError:
        return (
            "База знань ще не проіндексована. "
            "Запустіть переіндексацію: python ingest.py"
        )
    except Exception as exc:
        return f"Помилка завантаження бази знань: {exc}"

    try:
        results = retriever.search(q)
    except Exception as exc:
        return f"Помилка пошуку: {exc}"

    if not results:
        return (
            f"У базі знань не знайдено інформації за запитом «{q}». "
            "Спробуйте переформулювати запит або використайте lookup_tariff_telekom."
        )

    lines = [f"Знайдено {len(results)} релевантних фрагментів з бази знань:"]
    for i, chunk in enumerate(results[:4], 1):
        source = chunk.get("filename", "?")
        text   = chunk.get("text", "").strip()
        lines.append(f"\n[Фрагмент {i} | {source}]")
        lines.append(text[:600])
    return "\n".join(lines)



@tool
def lookup_tariff_telekom(tariff_name: str) -> str:
    """Look up conditions for a tariff plan on the operator's business tariff archive.

    If the tariff is NOT found on the website, it means the conditions are
    individual (договірні / індивідуальні умови).
    Always call this tool when the customer asks about their tariff conditions.
    """
    name = str(tariff_name or "").strip()
    if not name:
        return "Назву тарифу не вказано."

    # --- Step 1: try fetching the tariff archive page directly ---
    page_text = ""
    try:
        downloaded = trafilatura.fetch_url(
            _TELEKOM_TARIFF_URL,
            timeout=settings.url_fetch_timeout_sec,
        )
        if downloaded:
            extracted = trafilatura.extract(downloaded)
            if extracted:
                page_text = extracted
    except Exception:
        pass

    if page_text:
        name_lower = name.lower()
        # Find the relevant section around the tariff name
        text_lower = page_text.lower()
        idx = text_lower.find(name_lower)
        if idx != -1:
            start = max(0, idx - 200)
            end   = min(len(page_text), idx + 1200)
            snippet = page_text[start:end].strip()
            return (
                f"Знайдено тариф «{name}» на сайті оператора:\n\n"
                f"{snippet}\n\n"
                f"Джерело: {_TELEKOM_TARIFF_URL}"
            )

    # --- Step 2: DuckDuckGo fallback search ---
    fallback_query = f'"{name}" тариф умови Telekom'
    try:
        results = list(
            DDGS(timeout=settings.url_fetch_timeout_sec).text(
                fallback_query,
                max_results=3,
            )
        )
        if results:
            lines = [f"Тариф «{name}» не знайдено напряму на сторінці архіву."]
            lines.append("Знайдені посилання через пошук:")
            for r in results:
                title   = r.get("title", "")
                url     = r.get("href", r.get("url", ""))
                snippet = r.get("body", r.get("snippet", ""))
                lines.append(f"\n• {title}\n  {url}\n  {snippet[:300]}")
            return "\n".join(lines)
    except Exception:
        pass

    # --- Step 3: not found → individual conditions ---
    return (
        f"Тариф «{name}» не знайдено на сторінці архіву тарифів Telekom. "
        "Це означає, що умови тарифу є "
        "індивідуальними (договірними). Для отримання деталей зверніться "
        "до свого менеджера або служби підтримки Telekom."
    )


# ---------------------------------------------------------------------------
# Tool 3 — Escalation e-mail
# ---------------------------------------------------------------------------

def _build_email(
    sender: str,
    recipient: str,
    subject: str,
    summary: str,
    transcript: str,
    customer_info: str,
) -> MIMEMultipart:
    """Assemble a MIME email with a summary body and full transcript attachment."""
    msg = MIMEMultipart()
    msg["From"]    = sender
    msg["To"]      = recipient
    msg["Subject"] = subject

    today = date.today().isoformat()
    body = (
        f"Дата: {today}\n\n"
        f"Інформація про клієнта:\n{customer_info}\n\n"
        f"Резюме звернення:\n{summary}\n\n"
        "Повна транскрипція розмови додана як вкладення."
    )
    msg.attach(MIMEText(body, "plain", "utf-8"))

    # Attach the full transcript as a text file
    attachment = MIMEBase("application", "octet-stream")
    attachment.set_payload(transcript.encode("utf-8"))
    encoders.encode_base64(attachment)
    attachment.add_header(
        "Content-Disposition",
        "attachment",
        filename=f"chat_transcript_{today}.txt",
    )
    msg.attach(attachment)
    return msg


@tool
def send_escalation_email(
    target: str,
    subject: str,
    summary: str,
    session_id: str,
    customer_info: str = "",
    config: Annotated[RunnableConfig, InjectedToolArg] = None,
) -> str:
    """Send an escalation e-mail to the support or sales team.

    Args:
        target:        'support' — for unresolved complaints (customer not satisfied).
                       'sales'   — for tariff change requests or contract review.
        subject:       Brief issue description + client name (e-mail subject line).
        summary:       3–5 sentence summary of the issue and what was discussed.
        session_id:    Current session ID (used to fetch the full chat transcript from DB).
        customer_info: Client name, phone, account number, tariff — in one string.

    The full chat transcript is fetched from the local DB and attached automatically.
    The escalation is also recorded in the local database.
    """
    target = str(target or "").strip().lower()
    if target not in {"support", "sales"}:
        return "Невірний target. Вкажіть 'support' або 'sales'."

    # Read the authoritative db_session_id injected via LangGraph configurable.
    # This is set per-request in api.py and per-session in support_main.py,
    # so it is always correct — no race condition, no LLM hallucination risk.
    if config:
        db_sid = (config.get("configurable") or {}).get("db_session_id", "")
        if db_sid:
            session_id = db_sid

    smtp_host     = settings.escalation_smtp_host
    smtp_port     = settings.escalation_smtp_port
    smtp_user     = settings.escalation_smtp_sender
    smtp_password = settings.escalation_smtp_password.get_secret_value()

    recipient = (
        settings.escalation_email_support
        if target == "support"
        else settings.escalation_email_sales
    )

    # Fetch full transcript from DB
    transcript = format_chat_transcript(session_id)

    # Always persist to DB regardless of SMTP availability
    save_escalation(
        session_id=session_id,
        customer_info=customer_info,
        target=target,
        email_subject=subject,
        email_summary=summary,
        email_full_text=transcript,
    )
    save_identified_customer(
        session_id=session_id,
        client_name=customer_info[:200],
    )

    # If SMTP credentials not configured → log-only mode
    if not smtp_user or not smtp_password or not recipient:
        return (
            f"Ескалація збережена в базі даних (target='{target}').\n"
            "⚠️  SMTP не налаштований — лист не надіслано. "
            "Заповніть ESCALATION_SMTP_* та ESCALATION_EMAIL_* у файлі .env."
        )

    try:
        email_msg = _build_email(
            sender=smtp_user,
            recipient=recipient,
            subject=subject,
            summary=summary,
            transcript=transcript,
            customer_info=customer_info,
        )
        context = ssl.create_default_context()
        with smtplib.SMTP(smtp_host, smtp_port, timeout=30) as server:
            server.ehlo()
            server.starttls(context=context)
            server.login(smtp_user, smtp_password)
            server.sendmail(smtp_user, recipient, email_msg.as_string())

        team_label = "службу підтримки" if target == "support" else "відділ продажів"
        return (
            f"✅ Ескалацію надіслано до {team_label} ({recipient}).\n"
            f"Тема: {subject}\n"
            "Повна транскрипція прикріплена як вкладення.\n"
            "Запис збережено в базі даних."
        )
    except Exception as exc:
        return (
            f"⚠️  Не вдалося надіслати лист (target='{target}'): {exc}\n"
            "Ескалація збережена в базі даних."
        )
